import json
import os
import sys
import re
from graphviz import Digraph
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import argparse

def load_kanji_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read().strip()
        # Remove any trailing commas
        content = re.sub(r',\s*}', '}', content)
        content = re.sub(r',\s*]', ']', content)
        # Ensure the content is wrapped in curly braces
        if not content.startswith('{'):
            content = '{' + content
        if not content.endswith('}'):
            content = content + '}'
        # Add quotes around unquoted keys
        content = re.sub(r'(\s*)([^"\s:]+)(\s*:)', r'\1"\2"\3', content)
        try:
            print("Content being parsed:", content)
            return json.loads(content)
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            print("Content causing the error:")
            print(content)
            raise

def generate_kanji_diagram_recursive(kanji, kanji_data, dot, processed_kanji=None, level=0):
    if processed_kanji is None:
        processed_kanji = set()
    if kanji in processed_kanji or level > 2:  # Limit recursion depth to 2 levels
        return
    processed_kanji.add(kanji)
    
    fillcolor = '#FFFFFF' if level == 0 else '#99CCFF' if level == 1 else '#FFCCCC'
    dot.node(kanji, kanji, shape='circle', style='filled', fillcolor=fillcolor, 
             fontcolor='black', width='1', height='1', fontsize='24')
    
    if kanji in kanji_data:
        for in_kanji in kanji_data[kanji]['in']:
            if in_kanji in kanji_data:
                generate_kanji_diagram_recursive(in_kanji, kanji_data, dot, processed_kanji, level + 1)
            else:
                dot.node(in_kanji, in_kanji, shape='circle', style='filled', 
                         fillcolor='#FFCCCC', fontcolor='black')
            dot.edge(in_kanji, kanji)
        
        if level == 0:
            for out_kanji in kanji_data[kanji]['out'][:5]:
                dot.node(out_kanji, out_kanji, shape='circle', style='filled', 
                         fillcolor='#99CCFF', fontcolor='black')
                dot.edge(kanji, out_kanji)

def create_kanji_diagram(kanji, kanji_data, output_dir):
    dot = Digraph(comment=f'Kanji Diagram for {kanji}')
    dot.attr(rankdir='TB', size='12,12')
    generate_kanji_diagram_recursive(kanji, kanji_data, dot)
    output_file = os.path.join(output_dir, f"kanji_diagram_{kanji}")
    dot.render(output_file, format='png', cleanup=True)
    return f"{output_file}.png"

def create_excel_with_kanji_data(kanji_data, output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kanji Data"

    # Set column headers
    headers = ["Kanji", "In", "Out", "Diagram"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Populate data
    for row, (kanji, data) in enumerate(kanji_data.items(), start=2):
        ws.cell(row=row, column=1, value=kanji)
        ws.cell(row=row, column=2, value=', '.join(data['in']))
        ws.cell(row=row, column=3, value=', '.join(data['out']))

        # Generate and add diagram
        diagram_path = create_kanji_diagram(kanji, kanji_data, output_dir)
        img = Image(diagram_path)
        img.width = 200
        img.height = 200
        ws.add_image(img, f'D{row}')

    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 30

    # Save the workbook
    excel_path = os.path.join(output_dir, "kanji_data.xlsx")
    wb.save(excel_path)
    print(f"Excel file saved as {excel_path}")

def main():
    parser = argparse.ArgumentParser(description="Generate Kanji Excel file with diagrams")
    parser.add_argument("input_file", help="Path to the input text file containing kanji data")
    parser.add_argument("-o", "--output_dir", default="kanji_output", help="Directory to store output files (default: kanji_output)")
    args = parser.parse_args()

    input_file = args.input_file
    output_dir = args.output_dir

    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found.")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    kanji_data = load_kanji_data(input_file)
    create_excel_with_kanji_data(kanji_data, output_dir)
    print(f"Processing complete. Check the '{output_dir}' directory for output files.")

if __name__ == "__main__":
    main()