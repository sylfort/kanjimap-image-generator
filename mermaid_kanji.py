import json
import os
import sys
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import argparse

def load_kanji_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read().strip()
        content = re.sub(r',\s*}', '}', content)
        content = re.sub(r',\s*]', ']', content)
        if not content.startswith('{'):
            content = '{' + content
        if not content.endswith('}'):
            content = content + '}'
        content = re.sub(r'(\s*)([^"\s:]+)(\s*:)', r'\1"\2"\3', content)
        try:
            return json.loads(content)
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            print("Content causing the error:")
            print(content)
            raise

def create_mermaid_diagram(kanji, kanji_data):
    diagram = "```mermaid\ngraph TD\n"
    diagram += f"    {kanji}(({kanji}))\n"
    
    for in_kanji in kanji_data[kanji]['in']:
        diagram += f"    {in_kanji}(({in_kanji})) --> {kanji}\n"
    
    for i, out_kanji in enumerate(kanji_data[kanji]['out'][:5]):  # Limit to 5 output kanji
        diagram += f"    {kanji} --> {out_kanji}(({out_kanji}))\n"
    
    diagram += "```"
    return diagram

def create_excel_with_kanji_data(kanji_data, output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kanji Data"

    headers = ["Kanji", "In", "Out", "Mermaid Diagram"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row, (kanji, data) in enumerate(kanji_data.items(), start=2):
        ws.cell(row=row, column=1, value=kanji)
        ws.cell(row=row, column=2, value=', '.join(data['in']))
        ws.cell(row=row, column=3, value=', '.join(data['out']))
        
        mermaid_diagram = create_mermaid_diagram(kanji, kanji_data)
        ws.cell(row=row, column=4, value=mermaid_diagram)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 50

    excel_path = os.path.join(output_dir, "kanji_data_mermaid.xlsx")
    wb.save(excel_path)
    print(f"Excel file saved as {excel_path}")

def main():
    parser = argparse.ArgumentParser(description="Generate Kanji Excel file with Mermaid diagrams")
    parser.add_argument("input_file", help="Path to the input text file containing kanji data")
    parser.add_argument("-o", "--output_dir", default="kanji_output", help="Directory to store output files (default: kanji_output)")
    args = parser.parse_args()

    input_file = args.input_file
    output_dir = args.output_dir

    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found.")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    kanji_data = load_kanji_data(input_file)
    create_excel_with_kanji_data(kanji_data, output_dir)
    print(f"Processing complete. Check the '{output_dir}' directory for output files.")

if __name__ == "__main__":
    main()